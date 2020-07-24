# Reading an excel file using Python
import bank_account
import investment
import xlrd
import openpyxl as op
from openpyxl.styles import Alignment
import sys

# sets up the format of the summary sheet
# is only called once, if the file didn't exist when called
# formats to desired format for each inserted cell
# defaults initial values to N/A or 0, depending on which section, bankAccount or Invest
def setUpSummary(summarySheet):
    bankSummarySetUp(summarySheet)
    investmentSetUp(summarySheet)


# set up the archive sheet
def setUpArchive(archiveSheet):
    archiveSheet['A1'] = "Bank Archive"
    archiveSheet['B1'] = "Checking"
    archiveSheet['C1'] = "Savings"
    archiveSheet['D1'] = "Money Market"
    archiveSheet['E1'] = "Total"

# fit the cells of the sheet parameter ws
def fitCells(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(horizontal="center")
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+6


# basically the start of "main()"

# argument handling
if (len(sys.argv) == 1):
    print("Usage:\npython finance_analyzer.py <inputFile.xlsx> <summaryFile.xlsx>"), exit()
elif (len(sys.argv) != 3):
    print("Too many arguments.\nUsage:\npython finance_analyzer.py <inputFile.xlsx> <summaryFile.xlsx>"), exit()

# Give the location of the file
loc = (sys.argv[1])

# To open Workbook
try:
    wb = xlrd.open_workbook(loc)
except:
    print("Input file does not exist. Please create the well-formatted input file and use it."), exit()
sheet = wb.sheet_by_index(0)

bankPresent = False
investmentPresent = False

# Loop through columns
for i in range(sheet.nrows):
    # creates the account information in the BankAccount class
    if (sheet.cell_value(i,0) == "Bank Account"):
        bankPresent = True
        account = bank_account.BankAccount(sheet.cell_value(i+1, 1), sheet.cell_value(i+2, 1), sheet.cell_value(i+3, 1))
    elif (sheet.cell_value(i,0) == "Investment"):
        investmentPresent = True
        _investment = investment.Investment(sheet.cell_value(i+1, 1), sheet.cell_value(i+2, 1), sheet.cell_value(i+3, 1),
                                            sheet.cell_value(i+4, 1), sheet.cell_value(i+5, 1), sheet.cell_value(i+6, 1))


# open the excel workbook, if doesn't exist, create it and ensure 'Summary' is first sheet
try:
    writeWB = op.load_workbook(sys.argv[2])
    writeWB.title = sys.argv[2][0:-5]
    summarySheet = writeWB["Summary"]
    archiveSheet = writeWB["Bank Archive"]
    investArchive = writeWB["Investment Archive"]
except:
    writeWB = op.Workbook()
    writeWB.title = sys.argv[2][0:-5] # set the title without the .xlsx
    summarySheet = writeWB.create_sheet("Summary")
    archiveSheet = writeWB.create_sheet("Bank Archive")
    investArchive = writeWB.create_sheet("Investment Archive")
    if (writeWB.sheetnames[0] != "Summary"): # if not summary, delete first
        writeWB.remove(writeWB[writeWB.sheetnames[0]])
    setUpSummary(summarySheet)
    setUpArchive(archiveSheet)
    setUpInvestArchive(investArchive)



# call the various methods that handle writing to the excel files, only if they've
# been included in the inputfile
if (bankPresent == True):
    bank_account.allocateBankAccountValues(summarySheet, archiveSheet, account)
    bank_account.allocateSummaryBankAccount(summarySheet, archiveSheet, account)
    account.printAccount()
    print()
if (investmentPresent == True):
    investment.allocateInvestmentValues(summarySheet, investArchive, _investment)
    investment.allocateSummaryInvestment(summarySheet, _investment)
    _investment.printInformation()
    print()

fitCells(summarySheet)
fitCells(archiveSheet)
fitCells(investArchive)

# save the file
writeWB.save(writeWB.title + '.xlsx')
print("Done")
